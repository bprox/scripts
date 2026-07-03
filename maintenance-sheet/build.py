import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ---------- Color palette ----------
NAVY = "1F3864"
SLATE = "44546A"
LIGHTGRAY = "D9D9D9"
DUE_GREEN = "C6E0B4"
OVERDUE_RED = "F4B6B6"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
CATEGORY_FILL = PatternFill("solid", fgColor=SLATE)
HELPER_FILL = PatternFill("solid", fgColor=LIGHTGRAY)
DUE_FILL = PatternFill("solid", fgColor=DUE_GREEN)
OVERDUE_FILL = PatternFill("solid", fgColor=OVERDUE_RED)
WHITE_BOLD = Font(color="FFFFFF", bold=True, name="Arial")
BOLD = Font(bold=True, name="Arial")
NORMAL = Font(name="Arial")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# =====================================================================
# SHEET 1: 2026 Maintenance Tracker
# =====================================================================
ws = wb.active
ws.title = "2026 Tracker"
ws.sheet_view.showGridLines = False

# Column layout: A Task, B Every(Months), C Start Month, D:O months, P Notes
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 9
ws.column_dimensions["C"].width = 9
for col in range(4, 16):  # D..O
    ws.column_dimensions[get_column_letter(col)].width = 9.5
ws.column_dimensions["P"].width = 34

# ---- Title ----
ws.merge_cells("A1:P1")
ws["A1"] = "2026 Household Maintenance Tracker"
ws["A1"].font = Font(color="FFFFFF", bold=True, size=16, name="Arial")
ws["A1"].fill = HEADER_FILL
ws["A1"].alignment = CENTER
ws.row_dimensions[1].height = 26

ws.merge_cells("A2:P2")
ws["A2"] = "Split-level | Crawl space | Full HVAC | Well water w/ softener + chlorinate/dechlorinate + RO | 1.5 acres | 2 Subarus (Legacy & Forester) | Ohio"
ws["A2"].font = Font(italic=True, size=9, name="Arial", color="595959")
ws["A2"].alignment = CENTER
ws.row_dimensions[2].height = 16

# ---- Legend ----
legend_row = 3
ws.cell(row=legend_row, column=2, value="Legend:").font = BOLD
ws.cell(row=legend_row, column=3, value="DUE").fill = DUE_FILL
ws.cell(row=legend_row, column=3).alignment = CENTER
ws.cell(row=legend_row, column=3).font = NORMAL
ws.cell(row=legend_row, column=5, value="OVERDUE").fill = OVERDUE_FILL
ws.cell(row=legend_row, column=5).alignment = CENTER
ws.merge_cells(start_row=legend_row, start_column=5, end_row=legend_row, end_column=6)
ws.cell(row=legend_row, column=8, value="Type X when completed").font = Font(italic=True, size=9, name="Arial")
ws.merge_cells(start_row=legend_row, start_column=8, end_row=legend_row, end_column=11)

# ---- Header row ----
HEADER_ROW = 5
headers = ["Task", "Every\n(Months)", "Start\nMonth"]
ws.cell(row=HEADER_ROW, column=1, value=headers[0])
ws.cell(row=HEADER_ROW, column=2, value=headers[1])
ws.cell(row=HEADER_ROW, column=3, value=headers[2])

month_dates = [datetime.date(2026, m, 1) for m in range(1, 13)]
for i, d in enumerate(month_dates):
    c = ws.cell(row=HEADER_ROW, column=4 + i, value=d)
    c.number_format = "mmm-yy"

ws.cell(row=HEADER_ROW, column=16, value="Notes")

for col in range(1, 17):
    c = ws.cell(row=HEADER_ROW, column=col)
    c.font = WHITE_BOLD
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER
ws.row_dimensions[HEADER_ROW].height = 30

# =====================================================================
# Task data: (category_name, [(task, every_months, start_month, notes), ...])
# =====================================================================

categories = [
("MONTHLY", [
    ("Check water softener salt level", 1, 1, "Top off as needed"),
    ("Test sump pump (pour water into pit)", 1, 1, ""),
    ("Check furnace/HVAC air filter, replace if dirty", 1, 1, "Note filter size below"),
    ("Test smoke & CO detectors (button test)", 1, 1, ""),
    ("Clean range hood filter", 1, 1, ""),
    ("Check water softener / RO lines for leaks", 1, 1, ""),
    ("Check attic / whole-house fan operates freely", 1, 1, ""),
    ("Empty RO tank / check RO air gap", 1, 1, ""),
]),
("BI-MONTHLY (Every 2 Months)", [
    ("Clean dishwasher filter + run cleaner cycle", 2, 1, ""),
    ("Clean washing machine (tub-clean cycle)", 2, 1, ""),
    ("Run garbage disposal w/ ice + citrus, check for odor", 2, 1, ""),
    ("Check RO faucet filter indicator light", 2, 2, ""),
    ("Check tire pressure - Legacy", 2, 1, ""),
    ("Check tire pressure - Forester", 2, 2, ""),
]),
("QUARTERLY / EVERY 3-4 MONTHS", [
    ("Clean AC condenser coils (exterior unit)", 4, 3, ""),
    ("Deep clean garbage disposal & check seals", 4, 2, ""),
    ("Inspect roof from ground (binoculars)", 4, 1, "Limited roof access - visual only"),
    ("Check attic & crawl space for moisture/pests", 4, 2, ""),
    ("Clean oven", 4, 2, ""),
    ("Rotate tires - Legacy", 4, 1, ""),
    ("Rotate tires - Forester", 4, 2, ""),
    ("Test sump pump battery backup", 4, 3, ""),
    ("Clean refrigerator coils", 4, 1, ""),
    ("Lube garage door track/hinges & pull-down ladder", 4, 3, ""),
    ("Inspect asphalt driveway for cracks", 4, 2, ""),
    ("Check softener brine tank & resin bed condition", 4, 1, ""),
]),
("LAWN & YARD CARE (Seasonal)", [
    ("Refresh mulch beds", 12, 4, "Spring"),
    ("1st fertilizer round (crabgrass pre-emergent)", 12, 4, "Early spring"),
    ("2nd fertilizer round", 12, 5, "Late spring"),
    ("Mid-summer weed control / spot treat", 12, 7, ""),
    ("3rd fertilizer round", 12, 9, "Early fall"),
    ("4th fertilizer round + core aerate & overseed", 12, 10, "Fall - best time in Ohio"),
    ("Final winterizer fertilizer round", 12, 11, "Late fall"),
    ("Trim shrubs/trees away from roofline & gutters", 12, 10, ""),
    ("Clean up & bag leaves", 12, 11, ""),
]),
("SEMI-ANNUAL (Spring & Fall)", [
    ("Professional HVAC tune-up (AC spring / furnace fall)", 6, 4, ""),
    ("Clean gutters", 6, 4, ""),
    ("Flush water heater", 6, 4, "Gas tank unit"),
    ("Service well water system (softener resin cleaner, chlorinate/dechlorinate media)", 6, 3, ""),
    ("Clean dryer vent (duct + exterior hood)", 6, 2, ""),
    ("Change whole-home air purifier filter", 6, 3, ""),
    ("Check crawl space vents & vapor barrier", 6, 4, ""),
    ("Winterize (fall) / reactivate (spring) outdoor spigots", 6, 4, ""),
    ("Replace furnace humidifier pad", 6, 10, "If equipped"),
]),
("ANNUAL", [
    ("Change smoke & CO detector batteries", 12, 1, ""),
    ("Professional well water test (bacteria/nitrates)", 12, 5, ""),
    ("Service/inspect water softener & chlorination system (pro)", 12, 5, ""),
    ("Professional roof inspection", 12, 6, "Multiple roofs, limited access - hire pro"),
    ("Inspect garage storage loft & pull-down ladder mechanism", 12, 7, ""),
    ("Inspect exterior caulking & weatherstripping", 12, 9, ""),
    ("Check attic insulation levels", 12, 10, ""),
    ("Inspect/recharge fire extinguishers", 12, 1, "Check gauge & expiration"),
    ("Winterize exterior (cover AC unit, drain/store hoses)", 12, 11, ""),
    ("Change toothbrush heads", 12, 6, ""),
]),
]

MONTH_HEADER_ROW = HEADER_ROW  # row where month dates live
row_idx = HEADER_ROW + 1

for cat_name, tasks in categories:
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=16)
    c = ws.cell(row=row_idx, column=1, value=cat_name)
    c.font = WHITE_BOLD
    c.fill = CATEGORY_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1

    for task, every, start, note in tasks:
        ws.cell(row=row_idx, column=1, value=task).alignment = LEFT
        ws.cell(row=row_idx, column=2, value=every).alignment = CENTER
        ws.cell(row=row_idx, column=3, value=start).alignment = CENTER
        ws.cell(row=row_idx, column=2).fill = HELPER_FILL
        ws.cell(row=row_idx, column=3).fill = HELPER_FILL
        ws.cell(row=row_idx, column=16, value=note).alignment = LEFT
        for col in range(4, 16):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = CENTER
            cell.border = BORDER
        for col in [1, 2, 3, 16]:
            ws.cell(row=row_idx, column=col).border = BORDER
        ws.cell(row=row_idx, column=1).font = NORMAL
        ws.cell(row=row_idx, column=16).font = Font(italic=True, size=9, name="Arial", color="595959")
        row_idx += 1

LAST_DATA_ROW = row_idx - 1
FIRST_DATA_ROW = HEADER_ROW + 2  # first actual task row overall (approx; category rows interspersed, fine for CF range)

# ---- Conditional formatting across the whole grid ----
grid_range = f"D{HEADER_ROW+1}:O{LAST_DATA_ROW}"

overdue_formula = (
    f'AND($B{HEADER_ROW+1}<>"",MOD(MONTH(D${MONTH_HEADER_ROW})-$C{HEADER_ROW+1},$B{HEADER_ROW+1})=0,'
    f'D{HEADER_ROW+1}="",DATE(2026,MONTH(D${MONTH_HEADER_ROW}),1)<TODAY())'
)
due_formula = (
    f'AND($B{HEADER_ROW+1}<>"",MOD(MONTH(D${MONTH_HEADER_ROW})-$C{HEADER_ROW+1},$B{HEADER_ROW+1})=0)'
)

ws.conditional_formatting.add(
    grid_range,
    FormulaRule(formula=[overdue_formula], fill=OVERDUE_FILL, stopIfTrue=True),
)
ws.conditional_formatting.add(
    grid_range,
    FormulaRule(formula=[due_formula], fill=DUE_FILL, stopIfTrue=True),
)

ws.freeze_panes = f"D{HEADER_ROW+1}"

print("Sheet1 rows used:", LAST_DATA_ROW)
wb.save("Household_Maintenance_Tracker.xlsx")